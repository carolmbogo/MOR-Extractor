
from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import load_workbook


DATE_RE = re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}$")

KUB_PAGE1_FIELDS = [
    ("Date", 43.8), ("Rainfall (in)", 108.3), ("Average Flow (MGD)", 141.8),
    ("Maximum Flow (MGD)", 175.4), ("Minimum Flow (MGD)", 209.0),
    ("Bypass Hours", 242.6), ("Influent Temperature (°C)", 276.1),
    ("BOD Raw (mg/L)", 309.7), ("BOD Intermediate (mg/L)", 343.3),
    ("BOD Final (mg/L)", 376.8), ("BOD Reduction (%)", 410.4),
    ("Suspended Solids Raw (mg/L)", 444.0),
    ("Suspended Solids Intermediate (mg/L)", 477.6),
    ("Suspended Solids Final (mg/L)", 511.1),
    ("Suspended Solids Reduction (%)", 544.7),
    ("Set Solids Raw", 578.3), ("Set Solids Intermediate", 611.9),
    ("Set Solids Final", 645.4), ("DO (mg/L)", 679.0),
    ("pH Raw", 712.6), ("pH Intermediate", 746.1), ("pH Final", 779.7),
    ("NH3-N Influent (ppm)", 813.3), ("Grease Influent (ppm)", 846.9),
    ("Lbs Cl2", 880.4), ("Total N Influent (mg/L)", 914.0),
    ("Total P Influent (mg/L)", 947.6), ("E. coli", 981.2),
    ("Cl2 Residual", 1014.7),
    ("NH3-N Compliance Final Effluent (ppm)", 1048.3),
    ("Grease Final Effluent (ppm)", 1081.9), ("Effluent Flow (MGD)", 1115.4),
    ("Total N Effluent (mg/L)", 1149.0), ("Total P Effluent (mg/L)", 1182.6),
]

KUB_PAGE2_FIELDS = [
    ("Date", 43.8), ("Secondary System DO (ppm)", 164.3), ("MLSS (ppm)", 220.2),
    ("SVI", 276.2), ("30 Minute Set Solids", 332.2), ("Waste (GPD)", 388.1),
    ("Cadmium Influent (mg/L)", 444.1), ("Cadmium Effluent (mg/L)", 500.0),
    ("Chromium Influent (mg/L)", 556.0), ("Chromium Effluent (mg/L)", 611.9),
    ("Copper Influent (mg/L)", 667.9), ("Copper Effluent (mg/L)", 723.8),
    ("Nickel Influent (mg/L)", 779.8), ("Nickel Effluent (mg/L)", 835.7),
    ("Zinc Influent (mg/L)", 891.7), ("Zinc Effluent (mg/L)", 947.7),
    ("Silver Influent (mg/L)", 1003.6), ("Silver Effluent (mg/L)", 1059.6),
    ("Lead Influent (mg/L)", 1115.5), ("Lead Effluent (mg/L)", 1171.5),
]


@dataclass
class DetectedDataset:
    name: str
    source_name: str
    dataframe: pd.DataFrame
    confidence: str
    notes: list[str]


def clean_text(value) -> str:
    """Return clean text; missing Excel values always become blank."""
    if value is None:
        return ""
    try:
        missing = pd.isna(value)
        if isinstance(missing, bool) and missing:
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return re.sub(r"\s+", " ", text)


def parse_value(text: str):
    text = clean_text(text)
    if text == "":
        return None
    if text.startswith("<") or text.startswith(">"):
        return text
    candidate = text.replace(",", "")
    try:
        return float(candidate)
    except Exception:
        return text


def word_center(word) -> float:
    return (float(word["x0"]) + float(word["x1"])) / 2.0


def group_words_by_row(words, tolerance=2.2):
    ordered = sorted(words, key=lambda w: (float(w["top"]), float(w["x0"])))
    rows = []
    for w in ordered:
        top = float(w["top"])
        if not rows or abs(top - rows[-1]["top"]) > tolerance:
            rows.append({"top": top, "words": [w]})
        else:
            rows[-1]["words"].append(w)
            rows[-1]["top"] = sum(float(x["top"]) for x in rows[-1]["words"]) / len(rows[-1]["words"])
    for row in rows:
        row["words"] = sorted(row["words"], key=lambda w: float(w["x0"]))
    return rows


def row_has_date(row) -> bool:
    return any(DATE_RE.match(clean_text(w["text"])) for w in row["words"])


def extract_date_rows(page):
    words = page.extract_words(
        x_tolerance=1, y_tolerance=2, keep_blank_chars=False, use_text_flow=False
    )
    rows = group_words_by_row(words)
    return [r for r in rows if row_has_date(r)], rows


def nearest_field(x: float, field_defs, tolerance: float):
    name, anchor = min(field_defs, key=lambda item: abs(x - item[1]))
    return name if abs(x - anchor) <= tolerance else None


def extract_known_page(page, field_defs, tolerance=16.5) -> pd.DataFrame:
    date_rows, _ = extract_date_rows(page)
    records = []
    for row in date_rows:
        record = {name: None for name, _ in field_defs}
        for word in row["words"]:
            text = clean_text(word["text"])
            x = word_center(word)
            if DATE_RE.match(text):
                record["Date"] = pd.to_datetime(text, errors="coerce")
                continue
            field = nearest_field(x, field_defs, tolerance)
            if field and field != "Date":
                value = parse_value(text)
                if record[field] is None:
                    record[field] = value
                else:
                    record[field] = f"{record[field]} {text}"
        if record.get("Date") is not None and not pd.isna(record["Date"]):
            records.append(record)
    return pd.DataFrame(records, columns=[f[0] for f in field_defs])


def is_kub_fourth_creek(pdf) -> bool:
    try:
        text = (pdf.pages[0].extract_text() or "").lower()
    except Exception:
        return False
    required = [
        "report of operation of wastewater treatment plant",
        "influent flows", "5-day bod", "set solids", "final effluent parameters",
    ]
    return sum(term in text for term in required) >= 4


def extract_kub_pdf(data: bytes, source_name: str) -> list[DetectedDataset]:
    out = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        if not pdf.pages:
            return out

        p1 = extract_known_page(pdf.pages[0], KUB_PAGE1_FIELDS)
        if not p1.empty:
            out.append(DetectedDataset(
                name="Daily MOR - Page 1", source_name=source_name, dataframe=p1,
                confidence="High",
                notes=[
                    "Recognized Fourth Creek/KUB MOR layout.",
                    "Columns are assigned using PDF x-coordinates, not character spacing.",
                ],
            ))

        if len(pdf.pages) >= 2:
            p2 = extract_known_page(pdf.pages[1], KUB_PAGE2_FIELDS)
            if not p2.empty:
                out.append(DetectedDataset(
                    name="Secondary System - Page 2", source_name=source_name, dataframe=p2,
                    confidence="High",
                    notes=[
                        "Recognized Fourth Creek/KUB secondary-system table.",
                        "This DO is separate from the page-1 DO column.",
                    ],
                ))
    return out


def cluster_positions(values: list[float], tolerance=11.0):
    if not values:
        return []
    values = sorted(values)
    clusters = [[values[0]]]
    for x in values[1:]:
        center = sum(clusters[-1]) / len(clusters[-1])
        if abs(x - center) <= tolerance:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    return [sum(c) / len(c) for c in clusters]


def find_header_rows(all_rows, first_data_top, max_rows=4):
    prior = [r for r in all_rows if r["top"] < first_data_top - 1]
    return prior[-max_rows:]


def generic_header_for_anchor(anchor, header_rows, boundaries):
    left, right = boundaries
    pieces = []
    for row in header_rows:
        words = [clean_text(w["text"]) for w in row["words"] if left <= word_center(w) < right]
        if words:
            pieces.append(" ".join(words))
    if not pieces:
        return ""
    leaf = pieces[-1]
    generic_leafs = {
        "raw", "inter.", "inter", "final", "ppm", "mg/l", "mgd",
        "in %", "%", "hours", "inches", "total n", "total p"
    }
    if leaf.lower() in generic_leafs and len(pieces) >= 2:
        return f"{pieces[-2]} {leaf}".strip()
    return leaf


def make_unique(names):
    counts = {}
    out = []
    for i, n in enumerate(names, start=1):
        name = clean_text(n) or f"Detected Column {i}"
        counts[name] = counts.get(name, 0) + 1
        out.append(f"{name} ({counts[name]})" if counts[name] > 1 else name)
    return out


def extract_generic_pdf(data: bytes, source_name: str) -> list[DetectedDataset]:
    datasets = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            date_rows, all_rows = extract_date_rows(page)
            if len(date_rows) < 2:
                continue

            value_x = []
            date_anchor = None
            for row in date_rows:
                for w in row["words"]:
                    text = clean_text(w["text"])
                    x = word_center(w)
                    if DATE_RE.match(text):
                        if date_anchor is None:
                            date_anchor = x
                    else:
                        value_x.append(x)

            anchors = cluster_positions(value_x, tolerance=10.5)
            if not anchors:
                continue

            first_top = min(r["top"] for r in date_rows)
            header_rows = find_header_rows(all_rows, first_top, max_rows=4)
            full_anchors = ([date_anchor] if date_anchor is not None else []) + anchors
            full_anchors = sorted(full_anchors)

            labels = []
            for idx, anchor in enumerate(full_anchors):
                if date_anchor is not None and abs(anchor - date_anchor) < 2:
                    labels.append("Date")
                    continue
                left = -1e9 if idx == 0 else (full_anchors[idx - 1] + anchor) / 2
                right = 1e9 if idx == len(full_anchors) - 1 else (anchor + full_anchors[idx + 1]) / 2
                labels.append(generic_header_for_anchor(anchor, header_rows, (left, right)))
            labels = make_unique(labels)

            records = []
            for row in date_rows:
                record = {label: None for label in labels}
                for w in row["words"]:
                    text = clean_text(w["text"])
                    x = word_center(w)
                    if DATE_RE.match(text) and "Date" in record:
                        record["Date"] = pd.to_datetime(text, errors="coerce")
                        continue
                    idx = min(range(len(full_anchors)), key=lambda i: abs(x - full_anchors[i]))
                    if abs(x - full_anchors[idx]) <= 16.0:
                        label = labels[idx]
                        if label == "Date":
                            continue
                        value = parse_value(text)
                        if record[label] is None:
                            record[label] = value
                        else:
                            record[label] = f"{record[label]} {text}"
                if record.get("Date") is not None and not pd.isna(record["Date"]):
                    records.append(record)

            df = pd.DataFrame(records, columns=labels)
            if len(df) >= 2 and len(df.columns) >= 2:
                datasets.append(DetectedDataset(
                    name=f"Generic daily table - Page {page_no}",
                    source_name=source_name, dataframe=df,
                    confidence="Review required",
                    notes=[
                        "This file did not match a saved MOR template.",
                        "Columns were inferred from repeated PDF x-positions and nearby header text.",
                        "Confirm sample values before exporting.",
                    ],
                ))
    return datasets



# ----------------------------
# Scanned PDF / OCR logic
# ----------------------------

MONTH_LOOKUP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}


def get_pdf_page_info(data: bytes) -> dict:
    """
    Lightweight PDF inspection used by the Streamlit UI before extraction.
    A PDF is considered scanned/image-based when almost no embedded text is
    available on the first few pages.
    """
    import fitz

    doc = fitz.open(stream=data, filetype="pdf")
    page_count = len(doc)
    sample_pages = min(page_count, 3)
    text_chars = 0

    for i in range(sample_pages):
        try:
            text_chars += len((doc[i].get_text("text") or "").strip())
        except Exception:
            pass

    doc.close()

    return {
        "page_count": page_count,
        "is_scanned": text_chars < 80,
        "suggested_pages": [1] if page_count else [],
    }


def _ocr_page_words(page, zoom=2.0):
    """Render one PDF page and OCR it while preserving word coordinates."""
    import fitz
    import pytesseract
    from PIL import Image

    pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
    mode = "RGB" if pix.n == 3 else "RGBA"
    image = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
    if mode == "RGBA":
        image = image.convert("RGB")

    ocr = pytesseract.image_to_data(
        image,
        config="--psm 6",
        output_type=pytesseract.Output.DATAFRAME,
    )

    words = []
    for _, row in ocr.iterrows():
        text = clean_text(row.get("text"))
        if not text:
            continue

        try:
            conf = float(row.get("conf", -1))
        except Exception:
            conf = -1

        if conf < 10:
            continue

        left = float(row["left"])
        top = float(row["top"])
        width = float(row["width"])
        height = float(row["height"])

        # Ignore common table-line OCR artifacts.
        if text in {"|", "||", "_", "__", "___", "{", "}", "[", "]"}:
            continue

        words.append(
            {
                "text": text,
                "x0": left,
                "x1": left + width,
                "top": top,
                "bottom": top + height,
                "cx": left + width / 2.0,
                "cy": top + height / 2.0,
                "conf": conf,
            }
        )

    return words, pix.width, pix.height


def _infer_month_year(words, source_name=""):
    text = " ".join(w["text"] for w in sorted(words, key=lambda x: (x["top"], x["x0"]))).lower()

    for name, month in MONTH_LOOKUP.items():
        m = re.search(rf"\b{name}\s+(20\d{{2}})\b", text)
        if m:
            return month, int(m.group(1))

    # Abbreviated month fallback.
    for name, month in MONTH_LOOKUP.items():
        m = re.search(rf"\b{name[:3]}\.?\s+(20\d{{2}})\b", text)
        if m:
            return month, int(m.group(1))

    # Common MOR filename fallback, e.g. "... MOR 07 26.pdf".
    stem = Path(source_name).stem
    matches = re.findall(r"(?<!\d)(0?[1-9]|1[0-2])[ _-](\d{2}|20\d{2})(?!\d)", stem)
    if matches:
        mm, yy = matches[-1]
        year = int(yy)
        if year < 100:
            year += 2000
        return int(mm), year

    return None, None


def _daily_candidate(text):
    t = clean_text(text)
    if DATE_RE.match(t):
        dt = pd.to_datetime(t, errors="coerce")
        if not pd.isna(dt):
            return int(dt.day)
    if re.fullmatch(r"\d{1,2}", t):
        day = int(t)
        if 1 <= day <= 31:
            return day
    return None


def _fit_daily_row_centers(words, page_width):
    """
    Find the day/date column and fit a straight line through the daily rows.
    This is intentionally tolerant of missed OCR days and stray numeric OCR.
    """
    candidates = []
    for w in words:
        # Daily identifiers should live near the left edge of the MOR table.
        if w["cx"] > page_width * 0.125:
            continue
        day = _daily_candidate(w["text"])
        if day is not None:
            candidates.append((day, w["cy"], w["cx"], w["text"]))

    if len(candidates) < 4:
        return None

    best = None
    n = len(candidates)

    # Robust two-point model selection. A true MOR daily sequence has nearly
    # constant vertical spacing, while other numbers in the left margin do not.
    for i in range(n):
        d1, y1, _, _ = candidates[i]
        for j in range(i + 1, n):
            d2, y2, _, _ = candidates[j]
            if abs(d2 - d1) < 3:
                continue

            slope = (y2 - y1) / (d2 - d1)
            if not (8 <= slope <= 55):
                continue

            intercept = y1 - slope * d1
            tol = max(4.0, abs(slope) * 0.28)

            inliers = {}
            residual_sum = 0.0
            for day, y, x, raw in candidates:
                residual = abs(y - (intercept + slope * day))
                if residual <= tol:
                    if day not in inliers or residual < inliers[day][0]:
                        inliers[day] = (residual, y, x, raw)

            score = len(inliers)
            if score:
                residual_sum = sum(v[0] for v in inliers.values()) / score

            model = (score, -residual_sum, slope, intercept, inliers)
            if best is None or model[:2] > best[:2]:
                best = model

    if best is None or best[0] < 4:
        return None

    _, _, slope, intercept, inliers = best

    # Refit using all inliers for a cleaner row-spacing estimate.
    xs = []
    ys = []
    for day, (_, y, _, _) in inliers.items():
        xs.append(float(day))
        ys.append(float(y))

    if len(xs) >= 2:
        xbar = sum(xs) / len(xs)
        ybar = sum(ys) / len(ys)
        denom = sum((x - xbar) ** 2 for x in xs)
        if denom:
            slope = sum((x - xbar) * (y - ybar) for x, y in zip(xs, ys)) / denom
            intercept = ybar - slope * xbar

    return {
        "slope": slope,
        "intercept": intercept,
        "inliers": inliers,
        "candidate_count": len(candidates),
    }


def _cluster_supported_x(points, tolerance, min_support=2):
    """
    Cluster repeated x-centers and retain positions seen in multiple daily rows.
    points is a list of (x, row_number).
    """
    if not points:
        return []

    points = sorted(points, key=lambda t: t[0])
    clusters = []

    for x, row_num in points:
        if not clusters:
            clusters.append({"xs": [x], "rows": {row_num}})
            continue

        center = sum(clusters[-1]["xs"]) / len(clusters[-1]["xs"])
        if abs(x - center) <= tolerance:
            clusters[-1]["xs"].append(x)
            clusters[-1]["rows"].add(row_num)
        else:
            clusters.append({"xs": [x], "rows": {row_num}})

    kept = []
    for c in clusters:
        if len(c["rows"]) >= min_support:
            kept.append(sum(c["xs"]) / len(c["xs"]))

    return kept


def _group_words_into_lines(words, tolerance=7.0):
    rows = []
    for w in sorted(words, key=lambda x: (x["cy"], x["x0"])):
        if not rows or abs(w["cy"] - rows[-1]["cy"]) > tolerance:
            rows.append({"cy": w["cy"], "words": [w]})
        else:
            rows[-1]["words"].append(w)
            rows[-1]["cy"] = sum(x["cy"] for x in rows[-1]["words"]) / len(rows[-1]["words"])

    for row in rows:
        row["words"] = sorted(row["words"], key=lambda x: x["x0"])
    return rows


def _ocr_header_name(anchor_idx, anchors, header_words):
    anchor = anchors[anchor_idx]
    left = -1e9 if anchor_idx == 0 else (anchors[anchor_idx - 1] + anchor) / 2.0
    right = 1e9 if anchor_idx == len(anchors) - 1 else (anchor + anchors[anchor_idx + 1]) / 2.0

    lines = _group_words_into_lines(header_words, tolerance=8.0)
    pieces = []

    for line in lines:
        text = " ".join(
            clean_text(w["text"])
            for w in line["words"]
            if left <= w["cx"] < right and clean_text(w["text"])
        )
        text = clean_text(text)
        if text and text not in pieces:
            pieces.append(text)

    if not pieces:
        return ""

    # Keep the most useful lower-level header pieces while avoiding a huge title.
    useful = []
    for p in pieces:
        low = p.lower()
        if any(skip in low for skip in ("operations report", "permit", "city of", "report of operation")):
            continue
        useful.append(p)

    useful = useful[-4:] if useful else pieces[-3:]
    compact = compact_header_name(useful)
    return compact or clean_text(" ".join(useful))



def _looks_like_scanned_kub(words):
    text = " ".join(clean_text(w["text"]) for w in words).lower()
    score = sum(
        term in text
        for term in (
            "fourth creek",
            "report of operation",
            "5-day",
            "final effluent",
            "secondary system",
        )
    )
    return score >= 2


def _extract_scanned_known_kub(words, page_width, source_name, page_no, model):
    """
    Use the saved Fourth Creek/KUB column anchors even when the PDF page is a scan.
    The anchor positions are scaled to the rendered page width.
    """
    import calendar

    text = " ".join(clean_text(w["text"]) for w in words).lower()
    if "secondary system" in text:
        field_defs = KUB_PAGE2_FIELDS
        dataset_name = f"Secondary System - Scanned Page {page_no}"
    else:
        field_defs = KUB_PAGE1_FIELDS
        dataset_name = f"Daily MOR - Scanned Page {page_no}"

    month, year = _infer_month_year(words, source_name=source_name)
    if month and year:
        n_days = calendar.monthrange(year, month)[1]
    else:
        observed = sorted(model["inliers"])
        n_days = min(max(max(observed) if observed else 31, 28), 31)

    slope = model["slope"]
    intercept = model["intercept"]
    half_band = max(4.5, abs(slope) * 0.36)

    # KUB saved anchors are in the source page's 1224-point coordinate system.
    scale = page_width / 1224.0
    scaled_fields = [(name, x * scale) for name, x in field_defs]
    tolerance = 18.5 * scale

    records = []
    for day in range(1, n_days + 1):
        cy = intercept + slope * day
        record = {name: None for name, _ in field_defs}

        if month and year:
            record["Date"] = pd.Timestamp(year=year, month=month, day=day)

        row_words = [
            w for w in words
            if abs(w["cy"] - cy) <= half_band
        ]

        for w in row_words:
            text_value = clean_text(w["text"])
            if not text_value:
                continue

            x = w["cx"]
            field = nearest_field(x, scaled_fields, tolerance)

            if not field or field == "Date":
                continue

            value = parse_value(text_value)
            if record[field] is None:
                record[field] = value
            else:
                record[field] = clean_text(f"{record[field]} {text_value}")

        records.append(record)

    df = pd.DataFrame(records, columns=[name for name, _ in field_defs])

    keep = ["Date"]
    for c in df.columns:
        if c == "Date":
            continue
        s = df[c].fillna("").astype(str).str.strip()
        if s.ne("").any():
            keep.append(c)
    df = df[keep]

    return DetectedDataset(
        name=dataset_name,
        source_name=source_name,
        dataframe=df,
        confidence="OCR - High layout confidence",
        notes=[
            f"Scanned Fourth Creek/KUB MOR layout recognized on page {page_no}.",
            "Saved KUB column positions were scaled to the scanned page, while OCR supplied the cell values.",
            "OCR can still misread faint individual numbers or symbols, so verify sample values before export.",
            "Blank source cells remain blank.",
        ],
    )


def extract_scanned_page(page, page_no, source_name):
    """
    OCR one selected scanned page and reconstruct a daily table from repeated
    row spacing + repeated x positions.
    """
    import calendar

    words, page_width, page_height = _ocr_page_words(page, zoom=2.0)
    if not words:
        return None

    model = _fit_daily_row_centers(words, page_width)
    if model is None:
        return None

    if _looks_like_scanned_kub(words):
        return _extract_scanned_known_kub(
            words,
            page_width,
            source_name,
            page_no,
            model,
        )

    slope = model["slope"]
    intercept = model["intercept"]

    month, year = _infer_month_year(words, source_name=source_name)
    if month and year:
        number_of_days = calendar.monthrange(year, month)[1]
    else:
        observed_days = sorted(model["inliers"])
        number_of_days = max(observed_days) if observed_days else 31
        number_of_days = max(number_of_days, 28)
        number_of_days = min(number_of_days, 31)

    row_centers = {day: intercept + slope * day for day in range(1, number_of_days + 1)}
    half_band = max(4.5, abs(slope) * 0.36)

    # Gather words by fitted daily row, excluding the first/day column.
    row_words = {}
    x_points = []

    for day, cy in row_centers.items():
        selected = [
            w for w in words
            if abs(w["cy"] - cy) <= half_band
            and w["cx"] > page_width * 0.11
        ]
        row_words[day] = selected

        for w in selected:
            text = clean_text(w["text"])
            # Ignore obvious non-data line noise.
            if not text or text in {"-", "—", "=", "+"}:
                continue
            x_points.append((w["cx"], day))

    anchors = _cluster_supported_x(
        x_points,
        tolerance=max(9.0, page_width * 0.0075),
        min_support=2,
    )

    if len(anchors) < 2:
        return None

    first_row_y = row_centers[1]
    header_bottom = first_row_y - half_band
    header_top = max(0.0, first_row_y - abs(slope) * 6.5)

    header_words = [
        w for w in words
        if header_top <= w["cy"] < header_bottom
    ]

    labels = []
    for idx, anchor in enumerate(anchors):
        label = _ocr_header_name(idx, anchors, header_words)
        labels.append(label or f"Detected Column {idx + 1}")

    labels = make_unique(labels)

    records = []
    for day in range(1, number_of_days + 1):
        record = {}

        if month and year:
            record["Date"] = pd.Timestamp(year=year, month=month, day=day)
        else:
            record["Day"] = day

        words_this_row = sorted(row_words.get(day, []), key=lambda w: w["x0"])

        for w in words_this_row:
            x = w["cx"]
            idx = min(range(len(anchors)), key=lambda i: abs(x - anchors[i]))

            # Use midpoint ownership so a word cannot leak far into a wrong column.
            left = -1e9 if idx == 0 else (anchors[idx - 1] + anchors[idx]) / 2.0
            right = 1e9 if idx == len(anchors) - 1 else (anchors[idx] + anchors[idx + 1]) / 2.0
            if not (left <= x < right):
                continue

            label = labels[idx]
            text = clean_text(w["text"])
            value = parse_value(text)

            if label not in record or record[label] is None:
                record[label] = value
            else:
                # Join split OCR tokens such as "<" + "0.1".
                record[label] = clean_text(f"{record[label]} {text}")

        records.append(record)

    id_col = "Date" if month and year else "Day"
    df = pd.DataFrame(records)

    # Keep only columns with at least one nonblank value.
    keep = [id_col]
    for c in labels:
        if c not in df.columns:
            continue
        s = df[c].fillna("").astype(str).str.strip()
        if s.ne("").any():
            keep.append(c)

    df = df[keep]

    if len(df.columns) < 2:
        return None

    confidence = "OCR - Review required"
    notes = [
        f"Scanned PDF page {page_no} was read with OCR.",
        "Daily rows were reconstructed from the repeated day/date spacing on the page.",
        "OCR can misread faint numbers, symbols, or dense multi-row headers. Verify the sample values before export.",
        "Blank source cells remain blank in the extracted output.",
    ]

    return DetectedDataset(
        name=f"Scanned daily table - Page {page_no}",
        source_name=source_name,
        dataframe=df,
        confidence=confidence,
        notes=notes,
    )


def extract_scanned_pdf(data: bytes, source_name: str, selected_pages=None):
    """OCR only the pages the user selected."""
    import fitz

    doc = fitz.open(stream=data, filetype="pdf")
    page_count = len(doc)

    if selected_pages is None:
        selected_pages = [1]

    selected = []
    for p in selected_pages:
        try:
            p = int(p)
        except Exception:
            continue
        if 1 <= p <= page_count:
            selected.append(p)

    out = []
    for page_no in sorted(set(selected)):
        ds = extract_scanned_page(doc[page_no - 1], page_no, source_name)
        if ds:
            out.append(ds)

    doc.close()
    return out


def extract_pdf(data: bytes, source_name: str, selected_pages=None) -> list[DetectedDataset]:
    info = get_pdf_page_info(data)

    if info["is_scanned"]:
        return extract_scanned_pdf(
            data,
            source_name,
            selected_pages=selected_pages or info["suggested_pages"],
        )

    with pdfplumber.open(io.BytesIO(data)) as pdf:
        known = is_kub_fourth_creek(pdf)

    return extract_kub_pdf(data, source_name) if known else extract_generic_pdf(data, source_name)


# ----------------------------
# Excel multi-row header logic
# ----------------------------

UNIT_WORDS = {
    "mgd", "mgl", "mg/l", "ppm", "ppb", "inches", "inch", "in", "%",
    "cel", "deg c", "°c", "gpd", "lbs", "lb", "cfu", "ml", "mg"
}

GENERIC_WORDS = {
    "flow", "temp", "temperature", "raw", "final", "influent", "effluent",
    "daily", "max", "maximum", "min", "minimum", "avg", "average",
    "plant", "waste", "rain", "fall", "rainfall", "inf", "eff"
}


def normalize_unit(text: str) -> str:
    t = clean_text(text)
    low = t.lower()
    mapping = {
        "mgl": "mg/L",
        "mg/l": "mg/L",
        "mgd": "MGD",
        "gpd": "GPD",
        "ppm": "ppm",
        "ppb": "ppb",
        "cel": "°C",
        "deg c": "°C",
        "°c": "°C",
        "inches": "in",
        "inch": "in",
        "in": "in",
        "%": "%",
        "lbs": "lb",
        "lb": "lb",
    }
    return mapping.get(low, t)


def is_unit(text: str) -> bool:
    return clean_text(text).lower() in UNIT_WORDS


def score_header_row(row):
    vals = [clean_text(x) for x in row if clean_text(x)]
    if not vals:
        return -999
    numeric = 0
    for v in vals:
        try:
            float(v.replace(",", ""))
            numeric += 1
        except Exception:
            pass
    text_ratio = 1 - numeric / len(vals)
    uniqueness = len(set(vals)) / len(vals)
    keywords = (
        "date", "flow", "bod", "tss", "solid", "ph", "nh3", "ammonia",
        "grease", "do", "mlss", "svi", "chlor", "nitrogen", "phosph",
        "rain", "temp", "influent", "effluent"
    )
    keyword_score = min(sum(k in " ".join(vals).lower() for k in keywords) * 0.4, 2.0)
    return text_ratio * 2 + uniqueness + keyword_score


def apply_real_merged_headers(rows, header_top, header_bottom, merged_ranges):
    """
    Expand only actual Excel merged cells across the detected header band.
    Ordinary blank cells remain blank.
    """
    header_rows = [
        [clean_text(v) for v in row]
        for row in rows[header_top:header_bottom + 1]
    ]

    if not merged_ranges:
        return header_rows

    width = max((len(r) for r in header_rows), default=0)
    header_rows = [r + [""] * (width - len(r)) for r in header_rows]

    for r0, r1, c0, c1 in merged_ranges:
        if r1 <= header_top or r0 > header_bottom:
            continue

        source = ""
        if 0 <= r0 < len(rows) and 0 <= c0 < len(rows[r0]):
            source = clean_text(rows[r0][c0])

        if not source:
            continue

        start_r = max(r0, header_top)
        end_r = min(r1, header_bottom + 1)

        for rr in range(start_r, end_r):
            local_r = rr - header_top
            for cc in range(c0, min(c1, width)):
                header_rows[local_r][cc] = source

    return header_rows


SECTION_FAMILY_MARKERS = (
    "bod", "c.b.o.d", "cbod",
    "suspended solid", "settled solid",
    "ammonia nitrogen", "ammonia",
    "total nitrogen", "nitrogen",
    "total phosphorus", "phosphorus",
    "chlorine", "chlor",
    "dissolved oxygen",
    "e. coli", "fecal",
    "grease", "oil",
    "ph"
)


def is_section_family_heading(text: str) -> bool:
    """
    Return True for upper-level analyte/process family headings that should
    apply to multiple adjacent subcolumns, such as:
      5 DAY C.B.O.D.
      SUSPENDED SOLIDS
      SETTLED SOLIDS
      AMMONIA NITROGEN
    """
    t = clean_text(text)
    if not t:
        return False

    low = t.lower()

    # Avoid treating leaf labels such as "P-Chem INF" or units as section headings.
    if is_unit(t):
        return False

    return any(marker in low for marker in SECTION_FAMILY_MARKERS)


def propagate_section_family_headings(header_rows):
    """
    Carry recognized family headings horizontally across adjacent blank cells
    in the SAME upper header row until the next nonblank heading begins.

    This is intentionally narrower than a general forward-fill. Only known
    analyte/process family headings are propagated, so ordinary blank cells
    remain blank.

    Example:
        [ "5 DAY C.B.O.D.", "", "", "", "SUSPENDED SOLIDS", "", ... ]
    becomes:
        [ "5 DAY C.B.O.D.", "5 DAY C.B.O.D.", "5 DAY C.B.O.D.",
          "5 DAY C.B.O.D.", "SUSPENDED SOLIDS", "SUSPENDED SOLIDS", ... ]

    That lets all child columns inherit their common family parent.
    """
    if not header_rows:
        return header_rows

    width = max(len(r) for r in header_rows)
    rows = [list(r) + [""] * (width - len(r)) for r in header_rows]

    for r_idx, row in enumerate(rows):
        active_family = ""

        for c_idx, value in enumerate(row):
            val = clean_text(value)

            if val:
                if is_section_family_heading(val):
                    active_family = val
                else:
                    # A new explicit non-family heading ends the current section
                    # only when it looks like another upper-level block title.
                    # Leaf headings lower in the hierarchy usually occupy a
                    # different row and therefore do not interfere.
                    active_family = ""
                continue

            if active_family:
                row[c_idx] = active_family

        rows[r_idx] = row

    return rows

def header_path_for_column(header_rows, col_idx):
    pieces = []
    for row in header_rows:
        if col_idx >= len(row):
            continue
        val = clean_text(row[col_idx])
        if not val:
            continue
        if pieces and pieces[-1].lower() == val.lower():
            continue
        pieces.append(val)
    return pieces


def compact_header_name(path):
    """
    Build a concise field name from the merged parent heading,
    subheading, qualifier, and unit.
    """
    if not path:
        return ""

    units = [normalize_unit(x) for x in path if is_unit(x)]
    words = [clean_text(x) for x in path if clean_text(x) and not is_unit(x)]

    deduped = []
    for w in words:
        if not deduped or deduped[-1].lower() != w.lower():
            deduped.append(w)
    words = deduped

    if not words:
        return f"({units[-1]})" if units else ""

    structural_groups = {
        "plant influent", "plant effluent", "influent parameters",
        "final effluent parameters", "operations report", "daily", "date"
    }

    family_markers = (
        "suspended solid", "settled solid", "ammonia", "nitrogen",
        "phosph", "bod", "c.b.o.d", "cbod", "chlor", "grease",
        "dissolved oxygen", "fecal", "e. coli", "ph"
    )

    family_parent = next(
        (
            w for w in words
            if any(marker in w.lower() for marker in family_markers)
        ),
        None,
    )

    parent = family_parent or words[0]
    parent_low = parent.lower()
    preserve_parent = family_parent is not None

    working = [w for w in words if w.lower() not in structural_groups]
    if not working:
        working = words

    cleaned = []
    for w in working:
        if not cleaned or cleaned[-1].lower() != w.lower():
            cleaned.append(w)
    working = cleaned

    leaf = working[-3:] if len(working) > 3 else working[:]

    if preserve_parent and all(parent_low != w.lower() for w in leaf):
        leaf = [parent] + leaf

    name = " ".join(leaf).strip()

    replacements = {
        "MAX Flow": "Max Flow",
        "MIN Flow": "Min Flow",
        "Daily Daily Flow": "Daily Flow",
        "RAIN FALL": "Rainfall",
        "Rain Fall": "Rainfall",
        "RAW WASTE TEMP": "Raw Waste Temp",
        "FINAL EFF TEMP": "Final Eff Temp",
    }
    name = replacements.get(name, name)

    if units:
        unit = units[-1]
        if unit and f"({unit})" not in name:
            name = f"{name} ({unit})".strip()

    return name

def resolve_duplicate_headers(paths, provisional):
    """
    If two columns still have the same concise name, progressively prepend
    higher parent levels until the names become unique.
    """
    result = list(provisional)

    for _ in range(5):
        counts = {}
        for n in result:
            counts[n.lower()] = counts.get(n.lower(), 0) + 1

        changed = False
        for i, name in enumerate(result):
            if not name or counts.get(name.lower(), 0) <= 1:
                continue

            path = [x for x in paths[i] if x and not is_unit(x)]
            unit = next((normalize_unit(x) for x in reversed(paths[i]) if is_unit(x)), "")

            current_words = [x for x in re.sub(r"\s+\([^)]*\)$", "", name).split(" ") if x]
            current_low = " ".join(current_words).lower()

            # Find the nearest higher-level piece not already represented.
            chosen = None
            for parent in reversed(path[:-1]):
                plow = parent.lower()
                if plow not in current_low:
                    chosen = parent
                    break

            if chosen:
                base = re.sub(r"\s+\([^)]*\)$", "", name)
                new_name = f"{chosen} {base}".strip()
                if unit:
                    new_name += f" ({unit})"
                result[i] = new_name
                changed = True

        if not changed:
            break

    return make_unique(result)


def detect_excel_header_band(rows):
    """
    Find the likely top and bottom of a multi-row Excel header area.

    We score the first 15 populated rows, choose the strongest header-like row,
    then include nearby rows above/below that are mostly text or units.
    """
    max_scan = min(15, len(rows))
    if max_scan == 0:
        return 0, 0

    scores = [score_header_row(rows[i]) for i in range(max_scan)]
    anchor = max(range(max_scan), key=lambda i: scores[i])

    top = max(0, anchor - 4)
    bottom = min(len(rows) - 1, anchor + 4)

    # Stop before clearly data-like rows.
    final_bottom = anchor
    for i in range(anchor, bottom + 1):
        vals = [clean_text(v) for v in rows[i] if clean_text(v)]
        if not vals:
            final_bottom = i
            continue

        numeric = 0
        for v in vals:
            try:
                float(v.replace(",", ""))
                numeric += 1
            except Exception:
                pass

        numeric_ratio = numeric / len(vals)
        if i > anchor and numeric_ratio > 0.45:
            break
        final_bottom = i

    return top, final_bottom



SUMMARY_LABELS = {
    "tot", "tot.", "total",
    "avg", "avg.", "average",
    "max", "max.", "maximum",
    "min", "min.", "minimum",
    "mean", "median", "summary"
}


def is_daily_identifier(value) -> bool:
    """
    Accept a real Excel date, a date string, or a day-of-month value 1..31.
    MOR sheets commonly use just 1, 2, 3 ... 31 in the DATE column.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False

    if isinstance(value, (pd.Timestamp,)):
        return True

    # Python datetime/date values are recognized by pandas.
    if not isinstance(value, str):
        try:
            num = float(value)
            return num.is_integer() and 1 <= int(num) <= 31
        except Exception:
            pass

    s = clean_text(value)
    if not s:
        return False

    if DATE_RE.match(s):
        return True

    try:
        num = float(s)
        return num.is_integer() and 1 <= int(num) <= 31
    except Exception:
        return False


def row_contains_summary_label(row, look_at=3) -> bool:
    """
    Summary rows in MOR workbooks are normally labeled in the first few cells:
    TOT, AVG., MAX., MIN., etc.
    """
    for value in list(row)[:look_at]:
        label = clean_text(value).lower()
        if label in SUMMARY_LABELS:
            return True
    return False


def trim_excel_to_daily_rows(body_rows):
    """
    Keep only the monthly daily-data block and discard footer calculations,
    totals, averages, max/min rows, and unrelated tables below the MOR.

    Strategy:
      1. Find the first row whose first few cells contain a valid day/date.
      2. Once daily rows have started, keep rows with a day/date identifier.
      3. Stop immediately at explicit summary labels such as TOT/AVG/MAX/MIN.
      4. Also stop after the daily sequence ends, preventing lower tables
         (for example Total N / Total P summaries) from being imported.
    """
    if not body_rows:
        return body_rows

    kept = []
    started = False
    missed_after_start = 0

    for row in body_rows:
        cells = list(row)

        if started and row_contains_summary_label(cells):
            break

        # Find a day/date identifier in the first three cells. This handles
        # sheets with a blank spacer column before DATE.
        daily = any(is_daily_identifier(v) for v in cells[:3])

        if daily:
            kept.append(cells)
            started = True
            missed_after_start = 0
            continue

        if started:
            missed_after_start += 1

            # A single odd/blank row may occur inside some MORs, but two
            # consecutive non-daily rows means the monthly daily table ended.
            if missed_after_start >= 2:
                break

    return kept if kept else body_rows

def dataframe_from_rows(rows, sheet_name, source_name, merged_ranges=None):
    rows = [list(r) for r in rows]
    rows = [r for r in rows if any(clean_text(v) for v in r)]
    if len(rows) < 2:
        return None

    width = max(len(r) for r in rows)
    rows = [r + [None] * (width - len(r)) for r in rows]

    header_top, header_bottom = detect_excel_header_band(rows)
    header_rows = apply_real_merged_headers(
        rows, header_top, header_bottom, merged_ranges or []
    )

    # Some legacy MORs use centered headings across blank cells instead of
    # true Excel merged ranges. Propagate only recognized family headings
    # (e.g., 5 DAY C.B.O.D., SUSPENDED SOLIDS) across those blank siblings.
    header_rows = propagate_section_family_headings(header_rows)

    paths = [header_path_for_column(header_rows, c) for c in range(width)]
    headers = [compact_header_name(path) for path in paths]
    headers = resolve_duplicate_headers(paths, headers)

    # If everything went wrong, fall back to a single scored row.
    meaningful = sum(bool(clean_text(h)) and not h.startswith("Detected Column") for h in headers)
    if meaningful < 2:
        max_scan = min(12, len(rows))
        header_idx = max(range(max_scan), key=lambda i: score_header_row(rows[i]))
        headers = make_unique([clean_text(v) for v in rows[header_idx]])
        header_bottom = header_idx

    body = rows[header_bottom + 1:]

    # Keep only the actual daily MOR rows. This removes footer calculations
    # such as TOT / AVG / MAX / MIN and unrelated blocks below the month.
    body = trim_excel_to_daily_rows(body)

    normalized = [r[:len(headers)] for r in body]

    df = pd.DataFrame(normalized, columns=headers).dropna(how="all")

    # Remove columns that are completely blank.
    keep = []
    for c in df.columns:
        series = df[c].fillna("").astype(str).str.strip()
        if not series.ne("").any():
            continue
        keep.append(c)
    df = df[keep]

    if df.empty or len(df.columns) < 2:
        return None

    return DetectedDataset(
        name=f"Excel sheet: {sheet_name}",
        source_name=source_name,
        dataframe=df,
        confidence="High",
        notes=["Multi-row Excel headers were reconstructed from the vertical header hierarchy.", "Footer summary rows such as TOT, AVG, MAX, and MIN are excluded automatically."],
    )


def extract_excel(data: bytes, source_name: str, selected_sheets=None) -> list[DetectedDataset]:
    suffix = Path(source_name).suffix.lower()
    datasets = []

    if suffix == ".xls":
        import xlrd

        legacy_book = xlrd.open_workbook(file_contents=data, formatting_info=False)
        sheet_names = legacy_book.sheet_names()

        if selected_sheets is not None:
            wanted = set(selected_sheets)
            sheet_names = [s for s in sheet_names if s in wanted]

        for sheet_name in sheet_names:
            sh = legacy_book.sheet_by_name(sheet_name)
            rows = [sh.row_values(r) for r in range(sh.nrows)]
            merged_ranges = [
                (rlow, rhigh, clow, chigh)
                for (rlow, rhigh, clow, chigh) in sh.merged_cells
            ]

            ds = dataframe_from_rows(
                rows, sheet_name, source_name, merged_ranges=merged_ranges
            )
            if ds:
                datasets.append(ds)

        return datasets

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    wanted = set(selected_sheets) if selected_sheets is not None else None

    for ws in wb.worksheets:
        if wanted is not None and ws.title not in wanted:
            continue

        rows = [list(r) for r in ws.iter_rows(values_only=True)]

        merged_ranges = [
            (
                merged.min_row - 1,
                merged.max_row,
                merged.min_col - 1,
                merged.max_col,
            )
            for merged in ws.merged_cells.ranges
        ]

        ds = dataframe_from_rows(
            rows, ws.title, source_name, merged_ranges=merged_ranges
        )
        if ds:
            datasets.append(ds)

    return datasets



def get_excel_sheet_names(filename: str, data: bytes) -> list[str]:
    """Return worksheet names without parsing the sheet contents."""
    suffix = Path(filename).suffix.lower()

    if suffix == ".xls":
        book = pd.ExcelFile(io.BytesIO(data), engine="xlrd")
        return list(book.sheet_names)

    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        return list(wb.sheetnames)

    return []

def unpack_upload(filename: str, data: bytes):
    suffix = Path(filename).suffix.lower()

    if suffix in {".pdf", ".xls", ".xlsx", ".xlsm"}:
        return [(filename, data)]

    if suffix == ".zip":
        items = []
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            for info in z.infolist():
                if info.is_dir():
                    continue
                inner = Path(info.filename)
                ext = inner.suffix.lower()
                if ext in {".pdf", ".xls", ".xlsx", ".xlsm"}:
                    items.append((inner.name, z.read(info)))
        return items

    return []


def detect_file(
    filename: str,
    data: bytes,
    selected_sheets=None,
    selected_pdf_pages=None,
) -> list[DetectedDataset]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        return extract_pdf(data, filename, selected_pages=selected_pdf_pages)
    if suffix in {".xls", ".xlsx", ".xlsm"}:
        return extract_excel(data, filename, selected_sheets=selected_sheets)
    return []


def combine_same_named_datasets(datasets: list[DetectedDataset]):
    groups = {}
    for ds in datasets:
        key = (ds.name, tuple(ds.dataframe.columns), ds.confidence)
        groups.setdefault(key, []).append(ds)

    out = []
    for key, members in groups.items():
        if len(members) == 1:
            out.append(members[0])
            continue

        combined = pd.concat([m.dataframe for m in members], ignore_index=True)
        if "Date" in combined.columns:
            combined = combined.sort_values("Date", kind="stable").reset_index(drop=True)

        out.append(DetectedDataset(
            name=members[0].name,
            source_name=f"{len(members)} files",
            dataframe=combined,
            confidence=members[0].confidence,
            notes=members[0].notes + [f"Combined {len(members)} matching files."],
        ))
    return out
